import openpyxl

book = openpyxl.load_workbook('sub.xlsx')

active_sheet = book.active
print(active_sheet.max_column)
print(active_sheet.max_row)

count = 0

for cell_obj in list(active_sheet.columns)[2]:
    # for cell in column:
        print(cell_obj.value)

